#!/usr/bin/env python3
"""
自然语言测试用例信号匹配工具 — Flask Backend
Single-file backend using only: Flask, pandas, openpyxl, sqlite3 (all pre-installed).
"""
import hashlib
import json
import logging
import os
import re
import sqlite3
import tempfile
import time
import uuid
from datetime import datetime
from threading import Lock
from typing import Any, Dict, List, Optional, Tuple

from flask import Flask, jsonify, request, Response, send_from_directory

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = os.environ.get("DB_PATH", "./case_convert.db")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "./uploads")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "sk-placeholder")
DEEPSEEK_BASE_URL = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
DEEPSEEK_MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO")
FRONTEND_DIR = os.environ.get("FRONTEND_DIR", "../frontend")

logging.basicConfig(level=getattr(logging, LOG_LEVEL, logging.INFO),
                    format="%(asctime)s | %(levelname)-8s | %(message)s")
logger = logging.getLogger(__name__)

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__, static_folder=None)
app.config["JSON_ENSURE_ASCII"] = False

# ─────────────────────────────────────────────────────────────────────────────
# In-memory session store
# ─────────────────────────────────────────────────────────────────────────────
_sessions: Dict[str, Dict[str, Any]] = {}
_session_lock = Lock()

def set_session(sid: str, data: Dict[str, Any]):
    with _session_lock:
        _sessions[sid] = data

def get_session(sid: str) -> Optional[Dict[str, Any]]:
    with _session_lock:
        return _sessions.get(sid)

# ─────────────────────────────────────────────────────────────────────────────
# Database
# ─────────────────────────────────────────────────────────────────────────────
def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS uploaded_file (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_type TEXT, original_name TEXT, stored_path TEXT,
        file_ext TEXT, file_size INTEGER, file_hash TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS match_task (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_code TEXT UNIQUE, status TEXT,
        signal_file_id INTEGER, case_file_id INTEGER,
        model_name TEXT, model_base_url TEXT, temperature REAL,
        case_count INTEGER, matched_case_count INTEGER, unmatched_case_count INTEGER,
        started_at TEXT, finished_at TEXT, duration_ms INTEGER,
        error_message TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS signal_source (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, uploaded_file_id INTEGER,
        source_type TEXT, source_file_name TEXT, sheet_names_json TEXT,
        message_count INTEGER, signal_count INTEGER,
        normalized_data_json TEXT, signals_flatten_json TEXT,
        parse_status TEXT, parse_error_message TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS signal_item (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        signal_source_id INTEGER, message_id TEXT, message_id_hex TEXT,
        message_name TEXT, signal_name TEXT, signal_desc TEXT,
        values_json TEXT, unit TEXT, comment TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS case_batch (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, uploaded_file_id INTEGER,
        sheet_names_json TEXT, case_count INTEGER,
        column_mapping_json TEXT, parse_status TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS case_item (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        case_batch_id INTEGER, task_id INTEGER,
        row_index INTEGER, case_id TEXT, case_step TEXT, raw_row_json TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS case_semantics (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, case_item_id INTEGER,
        original_text TEXT, normalized_text TEXT, action TEXT,
        positions_json TEXT, expanded_steps_json TEXT,
        negative_patterns_json TEXT, enum_value_semantics_json TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS case_candidate_signal (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, case_item_id INTEGER,
        candidate_rank INTEGER, candidate_score REAL,
        signal_name TEXT, signal_desc TEXT, msg_id_hex TEXT,
        hit_reasons_json TEXT, values_json TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS prompt_record (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, case_item_id INTEGER,
        system_prompt TEXT, user_prompt TEXT,
        prompt_version TEXT, prompt_hash TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS llm_call_record (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, case_item_id INTEGER, prompt_record_id INTEGER,
        provider_name TEXT, model_name TEXT,
        response_text TEXT, response_json TEXT,
        http_status INTEGER, success INTEGER,
        error_message TEXT, latency_ms INTEGER, token_usage_json TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS case_match_result (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, case_item_id INTEGER, llm_call_record_id INTEGER,
        matched INTEGER, result_json TEXT, match_count INTEGER,
        info_str_summary TEXT, unmatched_reason TEXT,
        validation_status TEXT, validation_error_message TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS export_record (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER, export_file_name TEXT,
        export_status TEXT, created_at TEXT DEFAULT (datetime('now'))
    );
    """)
    conn.commit()
    conn.close()
    logger.info("Database initialized: %s", DB_PATH)

# ─────────────────────────────────────────────────────────────────────────────
# DBC Parser
# ─────────────────────────────────────────────────────────────────────────────
class _DBCParser:
    def __init__(self, dbc_file: str):
        self.messages: Dict[str, Any] = {}
        self._parse(dbc_file)

    def _parse(self, dbc_file: str):
        raw = open(dbc_file, "rb").read()
        for enc in ("utf-8-sig", "gbk", "gb2312", "utf-8", "latin-1"):
            try:
                content = raw.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            content = raw.decode("utf-8", errors="replace")
        content = content.replace("\r\n", "\n").replace("\r", "\n")

        val_map: Dict[str, Dict[str, Dict[str, str]]] = {}
        for m in re.finditer(r'VAL_\s+(\d+)\s+(\w+)\s+((?:\d+\s+"[^"]*"\s*)+);', content):
            mid, sn = m.group(1), m.group(2)
            val_map.setdefault(mid, {})[sn] = {
                p.group(1): p.group(2)
                for p in re.finditer(r'(\d+)\s+"([^"]*)"', m.group(3))
            }

        cm: Dict[str, Dict[str, str]] = {}
        for m in re.finditer(r'CM_\s+SG_\s+(\d+)\s+(\w+)\s+"([^"]*)"', content, re.DOTALL):
            cm.setdefault(m.group(1), {})[m.group(2)] = m.group(3).replace("\n", " ").strip()

        cyc: Dict[str, str] = {}
        for m in re.finditer(r'BA_\s+"GenMsgCycleTime"\s+BO_\s+(\d+)\s+(\d+)', content):
            cyc[m.group(1)] = m.group(2)

        bo_re = re.compile(r'BO_\s+(\d+)\s+(\w+)\s*:\s*(\d+)\s+(\w+)(.*?)(?=\nBO_\s|\Z)', re.DOTALL)
        sg_re = re.compile(
            r'SG_\s+(\w+)\s*(?:M|m\d+)?\s*:\s*\d+\|\d+@[01][+-]\s*'
            r'\([^)]+\)\s*\[[^\]]*\]\s*"([^"]*)"\s*([\w,\s]*)')

        for bo in bo_re.finditer(content):
            raw_id, mname, msize, sender, body = bo.groups()
            mid_int = int(raw_id) & 0x1FFFFFFF
            mid = str(mid_int)
            signals = {}
            for sg in sg_re.finditer(body):
                sn, unit, recvs_raw = sg.group(1), sg.group(2).strip(), sg.group(3).strip()
                recvs = [r.strip() for r in recvs_raw.split(",") if r.strip() and r.strip() != "Vector__XXX"]
                sd: Dict[str, Any] = {"signal_name": sn, "unit": unit or None, "receiver": recvs}
                if mid in val_map and sn in val_map[mid]:
                    sd["values"] = val_map[mid][sn]
                if mid in cm and sn in cm[mid]:
                    sd["comment"] = cm[mid][sn]
                signals[sn] = sd
            self.messages[mid] = {
                "message_id": mid, "message_id_hex": hex(mid_int),
                "message_name": mname, "message_size": msize,
                "node_name": sender if sender != "Vector__XXX" else None,
                "signals": signals, "cycle_time": cyc.get(raw_id),
            }

# ─────────────────────────────────────────────────────────────────────────────
# Signal Extractor
# ─────────────────────────────────────────────────────────────────────────────
_MSG_COL_ALIASES = {
    "message_id": ("message_id","msg_id","msgid","frame_id","can_id","bo","id",
        "message identifier","报文id","消息id","报文标识符","帧id"),
    "message_name": ("message_name","msg_name","frame_name","message","报文名","消息名","报文名称"),
    "message_size": ("message_size","dlc","length","frame_length","报文长度","字节数"),
    "node_name": ("node_name","sender","transmitter","发送节点","发送器","发送方","节点"),
}
_SIG_COL_ALIASES = {
    "signal_name": ("signal_name","sig_name","signal","name","信号名","信号名称"),
    "unit": ("unit","单位"),
    "default_value": ("default_value","initial_value","默认值","初始值"),
    "cycle_time": ("cycle_time","period","周期","发送周期"),
    "values": ("values","value_table","enum","enumeration","取值表","枚举值","值描述"),
    "comment": ("comment","description","desc","备注","描述"),
}

def _norm_col(c: str) -> str:
    return str(c).strip().lower().replace("\n","").replace("\r","").replace(" ","").replace("_","")

def _norm_text(v: Any) -> Optional[str]:
    if v is None: return None
    s = str(v).strip()
    return None if s.lower() in ("nan","none","") else s

def _norm_mid(v: Any) -> Optional[str]:
    t = _norm_text(v)
    if not t: return None
    low = t.lower()
    try:
        if low.startswith("0x"): return str(int(low, 16))
        if low.endswith("h"): return str(int(low[:-1], 16))
        n = float(low)
        return str(int(n)) if n.is_integer() else t
    except ValueError:
        return t

def _mid_hex(mid: Any) -> Optional[str]:
    try: return hex(int(str(mid)))
    except: return None

def _parse_val_map(text: str) -> Any:
    try:
        v = json.loads(text)
        if isinstance(v, dict): return v
    except: pass
    pairs = {}
    for part in text.replace("\n",";").split(";"):
        part = part.strip()
        if not part: continue
        if ":" in part: k, v = part.split(":", 1)
        elif "=" in part: k, v = part.split("=", 1)
        else: return text
        pairs[k.strip()] = v.strip()
    return pairs if pairs else text

def _build_alias_map() -> Dict[str, str]:
    am = {}
    for canonical, aliases in {**_MSG_COL_ALIASES, **_SIG_COL_ALIASES}.items():
        for a in aliases:
            am[_norm_col(a)] = canonical
    return am

_ALIAS_MAP = _build_alias_map()

def parse_signal_bytes(file_bytes: bytes, filename: str, sheet_name=None) -> Tuple[Dict, List]:
    import pandas as pd
    ext = os.path.splitext(filename)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes); tmp_path = tmp.name
    try:
        if ext == ".dbc":
            parser = _DBCParser(tmp_path)
            messages = parser.messages
            signal_count = sum(len(m.get("signals",{})) for m in messages.values())
            data = {"source_file": filename, "source_type": "dbc",
                    "message_count": len(messages), "signal_count": signal_count,
                    "sheet_names": [], "messages": messages}
        elif ext in (".xls",".xlsx",".xlsm"):
            data = _parse_excel_signal(tmp_path, filename, sheet_name, pd)
        else:
            raise ValueError(f"不支持的文件类型: {ext}")
        flat = _flatten_signals(data)
        return data, flat
    finally:
        try: os.unlink(tmp_path)
        except: pass

def _parse_excel_signal(path: str, filename: str, sheet_name, pd) -> Dict:
    workbook = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    if not isinstance(workbook, dict):
        workbook = {str(sheet_name) if sheet_name else "Sheet1": workbook}
    messages, parsed_sheets = {}, []
    for sheet, df in workbook.items():
        df = df.dropna(how="all")
        if df.empty: continue
        rename_map = {c: _ALIAS_MAP[_norm_col(c)] for c in df.columns if _norm_col(c) in _ALIAS_MAP}
        if "message_id" not in rename_map.values() or "signal_name" not in rename_map.values():
            continue
        df = df.rename(columns=rename_map).copy()
        nc = set(rename_map.values())
        for key in ("message_id","message_name","message_size","node_name"):
            if key in nc: df[key] = df[key].ffill()
        parsed_sheets.append(sheet)
        for _, row in df.iterrows():
            mid = _norm_mid(row.get("message_id"))
            sn = _norm_text(row.get("signal_name"))
            if not mid or not sn: continue
            msg = messages.setdefault(mid, {
                "message_id": mid, "message_id_hex": _mid_hex(mid),
                "message_name": _norm_text(row.get("message_name")),
                "message_size": _norm_text(row.get("message_size")),
                "node_name": _norm_text(row.get("node_name")), "signals": {},
            })
            sd = {"signal_name": sn}
            for field in ("unit","default_value","cycle_time","comment"):
                if field in df.columns:
                    v = _norm_text(row.get(field))
                    if v: sd[field] = v
            if "values" in df.columns:
                v = _norm_text(row.get("values"))
                if v: sd["values"] = _parse_val_map(v)
            msg["signals"][sn] = sd
    if not messages:
        raise ValueError("未在Excel中找到有效信号数据，请检查列名")
    signal_count = sum(len(m.get("signals",{})) for m in messages.values())
    return {"source_file": filename, "source_type": "excel",
            "message_count": len(messages), "signal_count": signal_count,
            "sheet_names": parsed_sheets, "messages": messages}

def _flatten_signals(data: Dict) -> List[Dict]:
    flat = []
    for msg in data.get("messages",{}).values():
        for sig in msg.get("signals",{}).values():
            flat.append({
                "msg_id": msg.get("message_id"),
                "msg_id_hex": msg.get("message_id_hex"),
                "message_name": msg.get("message_name"),
                "signal_name": sig.get("signal_name"),
                "signal_desc": sig.get("comment") or sig.get("signal_name"),
                "values": sig.get("values", {}),
                "unit": sig.get("unit"),
                "comment": sig.get("comment"),
            })
    return flat

# ─────────────────────────────────────────────────────────────────────────────
# Case Parser
# ─────────────────────────────────────────────────────────────────────────────
_CASE_ID_ALIASES = ["case_id","用例编号","测试用例id","用例id","编号","id","caseid","用例号"]
_CASE_STEP_ALIASES = ["case_step","测试步骤","步骤","用例描述","casestep","操作步骤","步骤描述","测试描述","description","step"]

def parse_case_bytes(file_bytes: bytes, filename: str, sheet_name=None) -> Dict:
    import pandas as pd
    ext = os.path.splitext(filename)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes); tmp_path = tmp.name
    try:
        workbook = pd.read_excel(tmp_path, sheet_name=sheet_name, dtype=str)
        if not isinstance(workbook, dict):
            workbook = {str(sheet_name) if sheet_name else "Sheet1": workbook}
        all_cases, sheet_names, col_map = [], [], {}
        for sheet, df in workbook.items():
            df = df.dropna(how="all")
            if df.empty: continue
            col_norm = {_norm_col(c): c for c in df.columns}
            cid_col = next((col_norm[_norm_col(a)] for a in _CASE_ID_ALIASES if _norm_col(a) in col_norm), None)
            step_col = next((col_norm[_norm_col(a)] for a in _CASE_STEP_ALIASES if _norm_col(a) in col_norm), None)
            if not cid_col or not step_col: continue
            sheet_names.append(sheet)
            col_map = {"case_id": cid_col, "case_step": step_col}
            for idx, row in df.iterrows():
                cid = _norm_text(row.get(cid_col))
                step = _norm_text(row.get(step_col))
                if not cid or not step: continue
                raw = {str(k): str(v) for k, v in row.items() if _norm_text(v)}
                all_cases.append({"row_index": int(idx)+2, "case_id": cid, "case_step": step, "raw_row": raw})
        if not all_cases:
            raise ValueError("未识别到有效测试用例数据，请检查列名（需包含用例编号和测试步骤相关列）")
        return {"sheet_names": sheet_names, "column_mapping": col_map, "case_count": len(all_cases), "cases": all_cases}
    finally:
        try: os.unlink(tmp_path)
        except: pass

# ─────────────────────────────────────────────────────────────────────────────
# Normalizer
# ─────────────────────────────────────────────────────────────────────────────
_OPEN_SYN = ["打开","开启","启动","接通","使能","激活","接入"]
_CLOSE_SYN = ["关闭","关掉","关断","断开","停止","禁用","去激活"]
_ACTION_NORM = {w:"打开" for w in _OPEN_SYN}
_ACTION_NORM.update({w:"关闭" for w in _CLOSE_SYN})

_RANGE_EXP = {
    "左侧":["左前","左后"], "左边":["左前","左后"],
    "右侧":["右前","右后"], "右边":["右前","右后"],
    "前方":["左前","右前"], "前侧":["左前","右前"], "前排":["左前","右前"],
    "后方":["左后","右后"], "后侧":["左后","右后"], "后排":["左后","右后"],
    "两侧":["左前","左后","右前","右后"], "四门":["左前","右前","左后","右后"],
    "全部":["左前","右前","左后","右后"], "所有":["左前","右前","左后","右后"],
    "整车":["左前","右前","左后","右后"], "全车":["左前","右前","左后","右后"],
}
_POS_ALIAS = {"主驾":["左前"], "副驾":["右前"], "左后":["后排左"], "右后":["后排右"]}
_NEG_MAP = {
    "未打开":"关闭","未开启":"关闭","未接通":"断开","未关闭":"打开",
    "未锁止":"解锁","未解锁":"锁止","未使能":"禁用","不是激活状态":"未激活",
    "不是开启状态":"关闭","不处于开启状态":"关闭","不是关闭状态":"打开","不处于关闭状态":"打开",
}
_ENUM_PATTERNS = [
    (re.compile(r"[一1]档|[一1]级"), "Level1"),
    (re.compile(r"[二2]档|[二2]级"), "Level2"),
    (re.compile(r"[三3]档|[三3]级"), "Level3"),
    (re.compile(r"高档|高级"), "High"),
    (re.compile(r"中档|中级|中等"), "Medium"),
    (re.compile(r"低档|低级"), "Low"),
    (re.compile(r"最大|最高|最强"), "__MAX__"),
    (re.compile(r"最小|最低|最弱"), "__MIN__"),
]

def normalize_case(text: str) -> Dict:
    neg_pats = [neg for neg in _NEG_MAP if neg in text]
    t = text
    for neg, pos in sorted(_NEG_MAP.items(), key=lambda x: -len(x[0])):
        t = t.replace(neg, pos)
    for syn, canon in _ACTION_NORM.items():
        t = t.replace(syn, canon)
    action = next((v for k, v in _ACTION_NORM.items() if k in text), "")
    positions = [rw for rw in _RANGE_EXP if rw in text] + [al for al in _POS_ALIAS if al in text]
    expanded = [t]
    for rw, locs in _RANGE_EXP.items():
        if rw in t:
            for loc in locs:
                expanded.append(t.replace(rw, loc))
    for alias, canon_list in _POS_ALIAS.items():
        if alias in t:
            for c in canon_list:
                expanded.append(t.replace(alias, c))
    expanded = list(dict.fromkeys(expanded))
    enum_sem = {m.group(): mapping for pat, mapping in _ENUM_PATTERNS for m in [pat.search(text)] if m}
    return {
        "original_text": text, "normalized_text": t, "action": action,
        "positions": positions, "expanded_steps": expanded,
        "negative_patterns": neg_pats, "enum_value_semantics": enum_sem,
    }

# ─────────────────────────────────────────────────────────────────────────────
# Signal Recall
# ─────────────────────────────────────────────────────────────────────────────
_MIN_KW_LEN = 2
_TOP_N = 25

# 汽车域信号名英文缩写/单词 → 中文语义（用于中文用例匹配英文信号名）
_EN2ZH: Dict[str, List[str]] = {
    # 位置/角色
    "drvr": ["主驾","驾驶","左前","驾驶员"],
    "drv":  ["主驾","驾驶","左前"],
    "pass": ["副驾","乘客","右前","前排右"],
    "pasgr":["副驾","乘客","右前"],
    "le":   ["左","左侧"],
    "ri":   ["右","右侧"],
    "re":   ["后排","后"],
    "frnt": ["前","前排"],
    "fr":   ["前","前排"],
    "rear": ["后","后排"],
    "lf":   ["左前"],
    "rf":   ["右前"],
    "lr":   ["左后","后左"],
    "rr":   ["右后","后右"],
    "left": ["左","左侧"],
    "right":["右","右侧"],
    "front":["前","前排"],
    "all":  ["全部","所有","整车"],
    # 功能
    "heat": ["加热","暖风","座椅加热"],
    "vent": ["通风","座椅通风"],
    "wind": ["通风","座椅通风"],
    "massg":["按摩","座椅按摩"],
    "massage":["按摩","座椅按摩"],
    "lck":  ["锁","锁止","车门锁"],
    "lock": ["锁","锁止","车门锁"],
    "unlk": ["解锁","开锁"],
    "door": ["车门","门"],
    "win":  ["车窗","窗"],
    "wdw":  ["车窗","窗"],
    "window":["车窗","窗"],
    "wiper":["雨刮","雨刮器"],
    "wipr": ["雨刮","雨刮器"],
    "lamp": ["灯","车灯"],
    "light":["灯","车灯","灯光"],
    "beam": ["远光","灯"],
    "hi":   ["远光","高"],
    "lo":   ["近光","低"],
    "turn": ["转向","转向灯"],
    "horn": ["喇叭","鸣笛"],
    "seat": ["座椅"],
    "steer":["方向盘","转向"],
    "swh":  ["方向盘"],
    "wheel":["车轮","方向盘"],
    "trunk":["后备箱","行李箱"],
    "hood": ["前机盖","发动机盖"],
    "sunrf":["天窗"],
    "sunroof":["天窗"],
    "mirr": ["后视镜","反光镜"],
    "mirror":["后视镜","反光镜"],
    "spd":  ["速度","车速"],
    "veh":  ["车辆","整车"],
    "chrg": ["充电"],
    "charge":["充电"],
    "ac":   ["空调"],
    "hvac": ["空调","暖通"],
    "fan":  ["风扇","鼓风机"],
    "temp": ["温度"],
    "pwr":  ["电源","功率"],
    "power":["电源","功率"],
    "eng":  ["发动机","引擎"],
    "mot":  ["电机","马达"],
    "brk":  ["制动","刹车"],
    "brake":["制动","刹车"],
    "park": ["驻车","停车"],
    "epb":  ["电子手刹","驻车制动"],
    "abs":  ["制动","防抱死"],
    "esp":  ["车身稳定","ESP"],
    "acc":  ["自适应巡航","ACC"],
    "lka":  ["车道保持"],
    "ldc":  ["车道偏离"],
    "fcw":  ["碰撞预警"],
    "aeb":  ["自动紧急制动"],
    "mem":  ["记忆","存储"],
    "info": ["信息"],
    "save": ["保存","存储","记忆"],
    "call": ["调用","召回"],
    "set":  ["设置","设定"],
    "sts":  ["状态"],
    "status":["状态"],
    "req":  ["请求","需求"],
    "fb":   ["反馈","回报"],
    "ctrl": ["控制"],
    "en":   ["使能","启用"],
    "dis":  ["禁用","关闭"],
    "on":   ["打开","开启","开"],
    "off":  ["关闭","关断","关"],
    "act":  ["激活","动作"],
    "inact":["未激活","关闭"],
    "fld":  ["折叠"],
    "fold": ["折叠"],
    "slide":["滑动","前后调节"],
    "hight":["高度"],
    "height":["高度"],
    "tilt": ["倾斜","靠背"],
    "recl": ["靠背","后仰"],
    "lumb": ["腰部","腰托"],
    "head": ["头枕"],
    "warn": ["警告","报警"],
    "alert":["警报","报警"],
    "err":  ["错误","故障"],
    "fault":["故障"],
    "batt": ["电池","蓄电池"],
    "vol":  ["电压","音量"],
    "curr": ["电流"],
    "soc":  ["电量","荷电状态"],
    "child":["儿童","童锁"],
    "safe": ["安全"],
    "srs":  ["安全气囊","SRS"],
    "airbag":["气囊","安全气囊"],
    "belt": ["安全带"],
    "buc":  ["安全带","扣合"],
    "occpn":["占用","乘员"],
    "crash":["碰撞"],
    "bcm":  ["车身控制","BCM"],
    "bdc":  ["车身控制"],
    "cdc":  ["座舱控制","CDC"],
    "vcu":  ["整车控制","VCU"],
    "mcu":  ["电机控制","MCU"],
    "bcu":  ["电池控制","BCU"],
    "dsm":  ["驾驶座椅模块","DSM"],
    "pdsm": ["副驾座椅模块"],
    "rdsm": ["后排座椅模块"],
    "tms":  ["热管理","TMS"],
    "network":["网络"],
    "remote":["远程","遥控"],
    "mode": ["模式"],
    "level":["等级","档位"],
    "max":  ["最大","最高"],
    "min":  ["最小","最低"],
    "antipin":["防夹"],
    "anti": ["防"],
    "pinch":["防夹"],
}

# 中文关键词 → 英文信号名词片段（用于从中文用例检索英文信号名）
_ZH2EN: Dict[str, List[str]] = {
    "主驾":   ["Drvr","Drv","Driver","Dr","Le","Lf"],
    "驾驶员": ["Drvr","Drv","Driver"],
    "副驾":   ["Pass","Pasgr","Passenger","Ri","Rf"],
    "乘客":   ["Pass","Pasgr","Passenger"],
    "左前":   ["Le","Lf","FrLe","FrntLe","Left","Drvr"],
    "右前":   ["Ri","Rf","FrRi","FrntRi","Right","Pass"],
    "左后":   ["ReLe","ReLe","RearLe","Lr"],
    "右后":   ["ReRi","ReRi","RearRi","Rr"],
    "前排":   ["Fr","Frnt","Front","Fr"],
    "后排":   ["Re","Rear","Rr"],
    "全部":   ["All"],
    "所有":   ["All"],
    "整车":   ["All","Veh"],
    "加热":   ["Heat"],
    "座椅加热":["Heat","SeatHeat"],
    "通风":   ["Wind","Vent"],
    "座椅通风":["Wind","Vent","SeatWind"],
    "按摩":   ["Massg","Massage"],
    "座椅按摩":["Massg","Massage","SeatMassg"],
    "座椅":   ["Seat"],
    "车门":   ["Door"],
    "车窗":   ["Win","Wdw","Window"],
    "车锁":   ["Lck","Lock"],
    "锁":     ["Lck","Lock"],
    "解锁":   ["Unlk","Unlock"],
    "雨刮":   ["Wipr","Wiper"],
    "天窗":   ["Sunrf","Sunroof"],
    "后备箱": ["Trunk"],
    "后视镜": ["Mirr","Mirror"],
    "方向盘": ["Swh","Steer","Wheel"],
    "转向灯": ["Turn"],
    "远光":   ["Hi","HighBeam","Beam"],
    "近光":   ["Lo","LowBeam"],
    "灯":     ["Lamp","Light"],
    "喇叭":   ["Horn"],
    "空调":   ["Ac","Hvac"],
    "充电":   ["Chrg","Charge"],
    "速度":   ["Spd","Speed"],
    "车速":   ["Spd","VehSpd"],
    "温度":   ["Temp"],
    "记忆":   ["Mem","Save","Info"],
    "保存":   ["Save","Mem"],
    "调用":   ["Call"],
    "折叠":   ["Fld","Fold"],
    "腰部":   ["Lumb"],
    "头枕":   ["Head"],
    "靠背":   ["Recl","Tilt"],
    "高度":   ["Hight","Height"],
    "安全带": ["Belt","Buc"],
    "安全气囊":["Srs","Airbag"],
    "防夹":   ["AntiPin","AntiPinch","Antipin"],
    "打开":   ["On","Open","Act","En"],
    "开启":   ["On","Open","Act","En"],
    "关闭":   ["Off","Close","Dis","Inact"],
    "设置":   ["Set"],
    "状态":   ["Sts","Status","St"],
    "请求":   ["Req"],
    "反馈":   ["Fb"],
    "控制":   ["Ctrl"],
    "模式":   ["Mode","Mod"],
    "等级":   ["Level"],
    "档位":   ["Level"],
    "制动":   ["Brk","Brake"],
    "刹车":   ["Brk","Brake"],
    "驻车":   ["Park","Epb"],
    "手刹":   ["Epb","Park"],
    "电机":   ["Mot"],
    "电池":   ["Batt"],
    "电量":   ["Soc"],
    "故障":   ["Fault","Err"],
    "报警":   ["Warn","Alert"],
    "儿童":   ["Child"],
    "童锁":   ["Child","ChildSafe"],
    "碰撞":   ["Crash","Fcw","Aeb"],
}

def _split_camel(name: str) -> List[str]:
    """拆解 CamelCase 信号名为词片段列表"""
    parts = re.findall(r'[A-Z][a-z0-9]*', name)
    return parts

def _signal_semantic_tags(sig_name: str) -> List[str]:
    """将信号名 CamelCase 拆分后，映射为中文语义标签"""
    parts = _split_camel(sig_name)
    tags = []
    for p in parts:
        pl = p.lower()
        if pl in _EN2ZH:
            tags.extend(_EN2ZH[pl])
    return list(dict.fromkeys(tags))

def _signal_searchable(sig: Dict) -> str:
    """构建信号的可搜索文本，包含信号名、CamelCase拆分词、中文语义标签、枚举值"""
    name = sig.get("signal_name", "")
    parts = _split_camel(name)
    tags = _signal_semantic_tags(name)
    desc = sig.get("signal_desc", "") or ""
    # 如果 signal_desc 含乱码（非ASCII且乱码字符占比高），忽略它
    if desc:
        printable = sum(1 for c in desc if ord(c) > 0x4e00 or c.isascii())
        if printable / max(len(desc), 1) < 0.3:
            desc = ""
    values_str = " ".join(v for v in sig.get("values", {}).values() if v and v.isascii())
    values_keys = " ".join(sig.get("values", {}).keys())
    msg_name = sig.get("message_name", "") or ""
    return " ".join(filter(None, [
        name,
        " ".join(parts),
        " ".join(tags),
        desc,
        values_str,
        values_keys,
        msg_name,
    ])).lower()

def _tokenize(text: str) -> List[str]:
    text = re.sub(r"[^\u4e00-\u9fa5a-zA-Z0-9_]", " ", text)
    tokens = [tok for tok in text.split() if len(tok) >= _MIN_KW_LEN]
    clean = re.sub(r"\s+","", re.sub(r"[^\u4e00-\u9fa5]"," ", text))
    for n in (2,3,4):
        tokens += [clean[i:i+n] for i in range(len(clean)-n+1)]
    return list(set(tokens))

def _expand_kw(text: str, sem: Dict) -> List[str]:
    kws = set(_tokenize(text))
    for s in sem.get("expanded_steps", []):
        kws.update(_tokenize(s))
    for rw, locs in _RANGE_EXP.items():
        if rw in text: kws.update(locs)
    for alias, canon in _POS_ALIAS.items():
        if alias in text: kws.update(canon)
    for word in _OPEN_SYN:
        if word in text: kws.update(_OPEN_SYN + ["打开"])
    for word in _CLOSE_SYN:
        if word in text: kws.update(_CLOSE_SYN + ["关闭"])
    ev = sem.get("enum_value_semantics", {})
    kws.update(ev.values())
    if "__MAX__" in ev.values(): kws.update(["Max","High","Level3","最高","最大"])
    if "__MIN__" in ev.values(): kws.update(["Min","Low","Level1","最低","最小"])
    for k in ev:
        kws.update(re.findall(r"\d+", k))
    # 中文词 → 英文信号名片段扩展（核心：让中文用例能匹配英文信号名）
    all_text = text + " " + " ".join(sem.get("expanded_steps", []))
    for zh, en_list in _ZH2EN.items():
        if zh in all_text:
            kws.update(en_list)
            kws.update([e.lower() for e in en_list])
    return [kw for kw in kws if len(str(kw)) >= _MIN_KW_LEN]

def recall_candidates(case_step: str, flat_signals: List[Dict], sem: Dict) -> List[Dict]:
    kws = _expand_kw(case_step, sem)
    kws_lower = [str(kw).lower() for kw in kws if len(str(kw)) >= _MIN_KW_LEN]
    scored = []
    for sig in flat_signals:
        searchable = _signal_searchable(sig)
        sig_name_lower = sig.get("signal_name", "").lower()
        sig_parts_lower = " ".join(_split_camel(sig.get("signal_name",""))).lower()
        sig_tags = " ".join(_signal_semantic_tags(sig.get("signal_name",""))).lower()
        score = 0.0
        hits = []
        for kl in kws_lower:
            if kl not in searchable:
                continue
            # 权重分级：信号名直接命中 > 信号名CamelCase片段命中 > 语义标签命中 > 其他
            if kl in sig_name_lower:
                w = 5.0
            elif kl in sig_parts_lower:
                w = 4.0
            elif kl in sig_tags:
                w = 3.0
            else:
                w = 1.0
            score += w
            hits.append(kl)
        if score > 0:
            scored.append({**sig, "score": round(score,2), "hit_reasons": list(dict.fromkeys(hits))[:15]})
    scored.sort(key=lambda x: -x["score"])
    return scored[:_TOP_N]

# ─────────────────────────────────────────────────────────────────────────────
# Prompt Builder
# ─────────────────────────────────────────────────────────────────────────────
_SYSTEM_PROMPT = """你是一个汽车CAN总线信号语义匹配专家。根据中文测试用例描述，从候选信号列表中精准匹配对应的CAN信号及目标枚举值。

【信号名命名规律】
信号名采用CamelCase英文缩写，每个词块含义如下（候选列表中已提供"语义"字段辅助理解）：
- 前缀：Cdc=座舱控制, Bcm/Bdc=车身控制, Vcu=整车控制, Dsm=驾驶座椅模块, Pdsm=副驾座椅, Rdsm=后排座椅
- 位置：Drvr/Dr=主驾/左前, Pass/Pasgr=副驾/右前, FrLe=左前, FrRi=右前, ReLe=左后, ReRi=右后, Re=后排, Fr/Frnt=前排
- 功能：Heat=加热, Wind/Vent=通风, Massg=按摩, Lck=锁, Door=车门, Win/Wdw=车窗, Seat=座椅
- 类型：Set=设置指令, Sts=状态反馈, Req=请求, Fb=反馈, Ctrl=控制

【语义映射规则】
- 主驾=驾驶员=左前 → Drvr/Dr/Le/FrLe
- 副驾=乘客=右前 → Pass/Pasgr/Ri/FrRi
- 后排左=左后 → ReLe/Lr; 后排右=右后 → ReRi/Rr
- 左侧=左前+左后, 右侧=右前+右后, 前排=左前+右前, 后排=左后+右后, 整车=全部四个位置
- 打开/开启/激活 → 枚举中On/Active/Enable/1等有效值
- 关闭/断开/去激活 → 枚举中Off/Inactive/Disable/0等关闭值
- N档=LevelN（1档=Level1, 2档=Level2, 3档=Level3）
- 最大/最高/最强 → 枚举中最高有效档位; 最小/最低 → 最低有效档位（排除NoReq/Reserved）
- 未打开=关闭, 未关闭=打开, 不是开启状态=关闭
- Set信号=主动设置指令, Sts信号=状态反馈（优先匹配Set类信号）

【严格规则】
1. 只能从候选信号中选择匹配项，严禁虚构不存在的信号名、报文ID或枚举值。
2. 一条用例可匹配多个信号（如"全部座椅加热"需匹配4个位置各自的信号）。
3. 枚举值必须返回枚举表中的键（数字字符串），不能返回枚举的描述文字。
4. 无充分证据时输出 matched=false，不允许猜测。
5. 输出必须是纯JSON，不含任何解释文字或代码块标记。

【输出格式】
{"case_id":"<ID>","case_step":"<步骤>","matched":true/false,"case_info":[{"signal_name":"<信号名>","msg_id":"<报文ID十六进制>","signal_desc":"<描述>","signal_val":"<枚举键>","info_str":"【<ID>, <信号名>, <值>】","match_reason":"<原因>"}],"unmatched_reason":null}

【示例1】用例：主驾座椅加热3档
候选：CdcDrvrSeatHeatSet(语义:主驾/座椅/加热/设置) 0x2D2 {0:No Req,1:Level1,2:Level2,3:Level3,6:Off}
输出：{"case_id":"tc_001","case_step":"主驾座椅加热3档","matched":true,"case_info":[{"signal_name":"CdcDrvrSeatHeatSet","msg_id":"0x2d2","signal_desc":"主驾座椅加热设置","signal_val":"3","info_str":"【0x2D2, CdcDrvrSeatHeatSet, 3】","match_reason":"Drvr=主驾，Heat=加热，3档=Level3，枚举键3"}],"unmatched_reason":null}

【示例2】用例：关闭全部座椅加热
候选：CdcDrvrSeatHeatSet(主驾加热设置)/CdcPassSeatHeatSet(副驾加热设置)/CdcReLeseatHeatSet(后左加热设置)/CdcReRiseatHeatSet(后右加热设置) 枚举{6:Off}
输出：{"case_id":"tc_002","case_step":"关闭全部座椅加热","matched":true,"case_info":[{"signal_name":"CdcDrvrSeatHeatSet","msg_id":"0x2d2","signal_desc":"主驾座椅加热设置","signal_val":"6","info_str":"【0x2D2, CdcDrvrSeatHeatSet, 6】","match_reason":"全部=四个位置，关闭=Off=枚举键6"},{"signal_name":"CdcPassSeatHeatSet","msg_id":"0x2d2","signal_desc":"副驾座椅加热设置","signal_val":"6","info_str":"【0x2D2, CdcPassSeatHeatSet, 6】","match_reason":"全部=四个位置，关闭=Off=枚举键6"}],"unmatched_reason":null}"""

def _fmt_cand_line(i: int, c: Dict) -> str:
    """格式化候选信号行，关键：用语义标签代替乱码的signal_desc"""
    name = c.get("signal_name", "")
    tags = _signal_semantic_tags(name)
    semantic = "/".join(tags[:6]) if tags else "—"
    vals = c.get("values", {})
    # 过滤掉 Reserved/NotUsed 等无意义枚举项，保留有信息量的
    useful_vals = {k: v for k, v in vals.items()
                   if v and not v.lower().startswith(("reserved", "notused", "not used", "not_used"))}
    vals_str = json.dumps(useful_vals, ensure_ascii=False) if useful_vals else (
        json.dumps(vals, ensure_ascii=False) if vals else "无枚举")
    msg_id = c.get("msg_id_hex") or c.get("msg_id", "—")
    return f"{i}. {name}(语义:{semantic}) | 报文ID:{msg_id} | 枚举:{vals_str}"

def build_prompts(case_id: str, case_step: str, sem: Dict, cands: List[Dict]) -> Dict:
    cand_lines = [_fmt_cand_line(i, c) for i, c in enumerate(cands, 1)]

    sem_notes = []
    if sem.get("negative_patterns"): sem_notes.append(f"否定式转换：{sem['negative_patterns']}")
    if sem.get("positions"): sem_notes.append(f"位置展开：{sem['positions']}")
    if sem.get("enum_value_semantics"): sem_notes.append(f"枚举语义：{sem['enum_value_semantics']}")
    expanded = sem.get("expanded_steps", [case_step])

    user_prompt = f"""【当前用例】
用例ID：{case_id}
原始步骤：{case_step}
归一化步骤：{sem.get('normalized_text', case_step)}

【语义说明】
{chr(10).join(f'  - {n}' for n in sem_notes) or '  - 无特殊转换'}

【展开子步骤】（位置/否定展开后的等价描述）
{chr(10).join(f'  - {s}' for s in expanded)}

【候选信号列表】（共{len(cands)}条，信号名括号内为语义解析）
{chr(10).join(cand_lines) if cand_lines else '  （无候选信号，请输出 matched=false）'}

请根据用例语义，从候选信号中精准匹配，严格按JSON格式输出，case_id固定为"{case_id}"。"""

    ph = hashlib.md5((_SYSTEM_PROMPT + user_prompt).encode()).hexdigest()
    return {"system_prompt": _SYSTEM_PROMPT, "user_prompt": user_prompt, "prompt_hash": ph, "version": "v2.0"}

# ─────────────────────────────────────────────────────────────────────────────
# LLM Client
# ─────────────────────────────────────────────────────────────────────────────
def _repair_json(text: str) -> Optional[Dict]:
    if not text: return None
    text = re.sub(r"```(?:json)?\s*", "", text)
    text = re.sub(r"```", "", text).strip()
    try: return json.loads(text)
    except: pass
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        try: return json.loads(m.group())
        except: pass
    return None

def call_llm(sys_prompt: str, user_prompt: str, model_cfg: Dict) -> Dict:
    try:
        import urllib.request, ssl
        api_key = DEEPSEEK_API_KEY
        base_url = model_cfg.get("base_url", DEEPSEEK_BASE_URL).rstrip("/")
        model = model_cfg.get("model", DEEPSEEK_MODEL)
        temperature = float(model_cfg.get("temperature", 0))
        timeout = int(model_cfg.get("timeout_seconds", 60))

        payload = json.dumps({
            "model": model, "temperature": temperature,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_prompt},
            ]
        }).encode()

        req = urllib.request.Request(
            f"{base_url}/v1/chat/completions",
            data=payload,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
            method="POST"
        )
        ctx = ssl.create_default_context()
        t0 = time.time()
        with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
            resp_data = json.loads(resp.read().decode())
        latency_ms = int((time.time()-t0)*1000)
        raw_text = resp_data["choices"][0]["message"]["content"]
        usage = resp_data.get("usage", {})
        parsed = _repair_json(raw_text)
        return {"response_text": raw_text, "response_json": parsed,
                "success": True, "error_message": None, "latency_ms": latency_ms,
                "http_status": 200, "token_usage": usage}
    except Exception as e:
        return {"response_text": "", "response_json": None, "success": False,
                "error_message": str(e), "latency_ms": 0, "http_status": 0, "token_usage": {}}

def validate_output(case_id: str, case_step: str, resp_json: Optional[Dict]) -> Dict:
    if resp_json is None:
        return {"case_id": case_id, "case_step": case_step, "matched": False,
                "case_info": [], "unmatched_reason": "模型返回无法解析为JSON", "validation_status": "parse_failed"}
    resp_json.setdefault("case_id", case_id)
    resp_json.setdefault("case_step", case_step)
    if "signals" in resp_json and "case_info" not in resp_json:
        resp_json["case_info"] = resp_json.pop("signals")
    matched = bool(resp_json.get("matched", False))
    case_info = []
    for item in resp_json.get("case_info", []):
        if isinstance(item, dict) and item.get("signal_name"):
            case_info.append({
                "signal_name": str(item.get("signal_name","")),
                "msg_id": str(item.get("msg_id","") or ""),
                "signal_desc": str(item.get("signal_desc","") or ""),
                "signal_val": str(item.get("signal_val","") or "") if item.get("signal_val") is not None else None,
                "info_str": str(item.get("info_str","") or ""),
                "match_reason": str(item.get("match_reason","") or ""),
            })
    return {
        "case_id": str(resp_json.get("case_id", case_id)),
        "case_step": str(resp_json.get("case_step", case_step)),
        "matched": matched, "case_info": case_info,
        "unmatched_reason": resp_json.get("unmatched_reason"),
        "validation_status": "ok",
    }

# ─────────────────────────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────────────────────────
def export_excel_bytes(orig_bytes: bytes, orig_name: str, results: List[Dict], case_id_col: str) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill
    result_map = {r["case_id"]: r for r in results}
    ext = os.path.splitext(orig_name)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
        f.write(orig_bytes); src = f.name
    out = src.replace(ext, f"_filled{ext}")
    try:
        wb = openpyxl.load_workbook(src)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        cid_idx = next((i+1 for i, h in enumerate(headers) if h and str(h).strip() == case_id_col), None)
        if not cid_idx:
            raise ValueError(f"列 '{case_id_col}' 未找到")
        nc = ws.max_column + 1
        for j, h in enumerate(["匹配状态","匹配信号汇总","未匹配原因","结构化结果JSON"]):
            ws.cell(1, nc+j, h)
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            cid = str(row[cid_idx-1].value or "").strip()
            r = result_map.get(cid)
            if not r: continue
            matched = r.get("matched", False)
            summary = " | ".join(c.get("info_str") or c.get("signal_name","") for c in r.get("case_info",[]))
            sc = ws.cell(row[0].row, nc, "成功" if matched else "失败")
            ws.cell(row[0].row, nc+1, summary)
            ws.cell(row[0].row, nc+2, r.get("unmatched_reason","") or "")
            ws.cell(row[0].row, nc+3, json.dumps(r, ensure_ascii=False))
            sc.fill = green if matched else red
        wb.save(out)
        with open(out, "rb") as f: return f.read()
    finally:
        for p in [src, out]:
            try: os.unlink(p)
            except: pass

# ─────────────────────────────────────────────────────────────────────────────
# Match Orchestration
# ─────────────────────────────────────────────────────────────────────────────
def run_match(task_id: int, flat_signals: List[Dict], cases: List[Dict], model_cfg: Dict, task_code: str) -> List[Dict]:
    results = []
    conn = get_db()
    conn.execute("UPDATE match_task SET status='matching', started_at=datetime('now') WHERE id=?", (task_id,))
    conn.commit()

    for case_data in cases:
        case_id = case_data["case_id"]
        case_step = case_data["case_step"]
        case_item_id = case_data.get("db_id")

        try:
            sem = normalize_case(case_step)
            conn.execute("""INSERT INTO case_semantics
                (task_id,case_item_id,original_text,normalized_text,action,
                 positions_json,expanded_steps_json,negative_patterns_json,enum_value_semantics_json)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (task_id, case_item_id, sem["original_text"], sem["normalized_text"], sem["action"],
                 json.dumps(sem["positions"]), json.dumps(sem["expanded_steps"]),
                 json.dumps(sem["negative_patterns"]), json.dumps(sem["enum_value_semantics"])))
            conn.commit()

            cands = recall_candidates(case_step, flat_signals, sem)
            for rank, c in enumerate(cands):
                conn.execute("""INSERT INTO case_candidate_signal
                    (task_id,case_item_id,candidate_rank,candidate_score,signal_name,signal_desc,msg_id_hex,hit_reasons_json,values_json)
                    VALUES (?,?,?,?,?,?,?,?,?)""",
                    (task_id, case_item_id, rank+1, c.get("score",0), c.get("signal_name"),
                     c.get("signal_desc"), c.get("msg_id_hex"), json.dumps(c.get("hit_reasons",[])),
                     json.dumps(c.get("values",{}))))
            conn.commit()

            prompts = build_prompts(case_id, case_step, sem, cands)
            cur = conn.execute("""INSERT INTO prompt_record
                (task_id,case_item_id,system_prompt,user_prompt,prompt_version,prompt_hash)
                VALUES (?,?,?,?,?,?)""",
                (task_id, case_item_id, prompts["system_prompt"], prompts["user_prompt"],
                 prompts["version"], prompts["prompt_hash"]))
            prompt_id = cur.lastrowid
            conn.commit()

            llm = call_llm(prompts["system_prompt"], prompts["user_prompt"], model_cfg)
            cur2 = conn.execute("""INSERT INTO llm_call_record
                (task_id,case_item_id,prompt_record_id,provider_name,model_name,
                 response_text,response_json,http_status,success,error_message,latency_ms,token_usage_json)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (task_id, case_item_id, prompt_id, "deepseek", model_cfg.get("model","deepseek-chat"),
                 llm["response_text"], json.dumps(llm["response_json"]) if llm["response_json"] else None,
                 llm["http_status"], int(llm["success"]), llm["error_message"], llm["latency_ms"],
                 json.dumps(llm["token_usage"])))
            llm_id = cur2.lastrowid
            conn.commit()

            validated = validate_output(case_id, case_step, llm.get("response_json"))
            info_sum = " | ".join(c.get("info_str","") for c in validated.get("case_info",[]) if c.get("info_str"))
            conn.execute("""INSERT INTO case_match_result
                (task_id,case_item_id,llm_call_record_id,matched,result_json,match_count,info_str_summary,unmatched_reason,validation_status)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (task_id, case_item_id, llm_id, int(validated.get("matched",False)),
                 json.dumps(validated, ensure_ascii=False), len(validated.get("case_info",[])),
                 info_sum, validated.get("unmatched_reason"), validated.get("validation_status")))
            conn.commit()

            results.append(validated)
            status_icon = "✅" if validated.get("matched") else "❌"
            logger.info(f"  {status_icon} [{case_id}] signals={len(validated.get('case_info',[]))}")

        except Exception as e:
            conn.rollback()
            logger.error(f"Error on case {case_id}: {e}")
            results.append({"case_id": case_id, "case_step": case_step, "matched": False,
                            "case_info": [], "unmatched_reason": f"处理异常: {str(e)[:200]}", "validation_status": "error"})

    matched = sum(1 for r in results if r.get("matched"))
    conn.execute("""UPDATE match_task SET status='success', finished_at=datetime('now'),
        matched_case_count=?, unmatched_case_count=?, case_count=? WHERE id=?""",
        (matched, len(results)-matched, len(results), task_id))
    conn.commit()
    conn.close()
    return results

# ─────────────────────────────────────────────────────────────────────────────
# Flask Routes
# ─────────────────────────────────────────────────────────────────────────────

def err(msg: str, code: int = 400):
    return jsonify({"detail": msg}), code

@app.route("/")
def index():
    return send_from_directory(os.path.abspath(FRONTEND_DIR), "index.html")

@app.route("/health")
def health():
    return jsonify({"status": "healthy"})

# ── Signals ────────────────────────────────────────────────────────────────────
@app.route("/api/signals/parse", methods=["POST"])
def api_parse_signal():
    f = request.files.get("file")
    if not f: return err("未收到文件")
    filename = f.filename or "upload"
    ext = filename.rsplit(".",1)[-1].lower() if "." in filename else ""
    if ext not in ("dbc","xls","xlsx","xlsm"):
        return err(f"不支持的文件类型: .{ext}")
    file_bytes = f.read()
    sheet_name = request.form.get("sheet_name") or None
    fhash = hashlib.md5(file_bytes).hexdigest()
    try:
        data, flat = parse_signal_bytes(file_bytes, filename, sheet_name)
    except Exception as e:
        return err(f"信号文件解析失败: {e}", 422)

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO uploaded_file (file_type,original_name,stored_path,file_ext,file_size,file_hash) VALUES (?,?,?,?,?,?)",
        ("signal", filename, "", ext, len(file_bytes), fhash))
    uf_id = cur.lastrowid
    cur2 = conn.execute(
        """INSERT INTO signal_source (uploaded_file_id,source_type,source_file_name,sheet_names_json,
           message_count,signal_count,normalized_data_json,signals_flatten_json,parse_status)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (uf_id, data["source_type"], filename, json.dumps(data.get("sheet_names",[])),
         data["message_count"], data["signal_count"],
         json.dumps(data, ensure_ascii=False), json.dumps(flat, ensure_ascii=False), "success"))
    ss_id = cur2.lastrowid
    for sig in flat:
        conn.execute(
            """INSERT INTO signal_item (signal_source_id,message_id,message_id_hex,message_name,signal_name,signal_desc,values_json,unit,comment)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (ss_id, sig.get("msg_id"), sig.get("msg_id_hex"), sig.get("message_name"),
             sig.get("signal_name"), sig.get("signal_desc"),
             json.dumps(sig.get("values",{})), sig.get("unit"), sig.get("comment")))
    conn.commit(); conn.close()

    sid = f"sig_{uuid.uuid4().hex}"
    set_session(sid, {"signal_source_id": ss_id, "uploaded_file_id": uf_id,
                      "flat_signals": flat, "file_bytes": file_bytes, "filename": filename})

    # 写缓存文件
    cache_dir = os.path.join(UPLOAD_DIR, "cache")
    os.makedirs(cache_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(filename)[0]
    cache_path = os.path.join(cache_dir, f"signal_{base}_{ts}_{fhash[:8]}.json")
    try:
        with open(cache_path, "w", encoding="utf-8") as cf:
            json.dump({"meta": {"session_id": sid, "filename": filename, "file_hash": fhash,
                                "source_type": data["source_type"], "message_count": data["message_count"],
                                "signal_count": data["signal_count"], "parsed_at": datetime.now().isoformat()},
                       "signals": flat}, cf, ensure_ascii=False, indent=2)
        logger.info("Signal cache written: %s", cache_path)
    except Exception as e:
        logger.warning("Failed to write signal cache: %s", e)

    return jsonify({
        "signal_session_id": sid,
        "source_type": data["source_type"],
        "message_count": data["message_count"],
        "signal_count": data["signal_count"],
        "sheet_names": data.get("sheet_names",[]),
        "signals_preview": [{"msg_id": s.get("msg_id_hex") or s.get("msg_id"),
                              "signal_name": s["signal_name"], "signal_desc": s.get("signal_desc","")}
                             for s in flat[:20]],
    })

# ── Cases ──────────────────────────────────────────────────────────────────────
@app.route("/api/cases/parse", methods=["POST"])
def api_parse_case():
    f = request.files.get("file")
    if not f: return err("未收到文件")
    filename = f.filename or "upload"
    ext = filename.rsplit(".",1)[-1].lower() if "." in filename else ""
    if ext not in ("xls","xlsx","xlsm"):
        return err("测试用例文件仅支持 Excel 格式")
    file_bytes = f.read()
    fhash = hashlib.md5(file_bytes).hexdigest()
    sheet_name = request.form.get("sheet_name") or None
    try:
        parsed = parse_case_bytes(file_bytes, filename, sheet_name)
    except Exception as e:
        return err(f"测试用例文件解析失败: {e}", 422)

    cases = parsed["cases"]
    col_map = parsed.get("column_mapping", {})
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO uploaded_file (file_type,original_name,stored_path,file_ext,file_size,file_hash) VALUES (?,?,?,?,?,?)",
        ("case", filename, "", ext, len(file_bytes), fhash))
    uf_id = cur.lastrowid
    cur2 = conn.execute(
        """INSERT INTO case_batch (uploaded_file_id,sheet_names_json,case_count,column_mapping_json,parse_status)
           VALUES (?,?,?,?,?)""",
        (uf_id, json.dumps(parsed.get("sheet_names",[])), len(cases),
         json.dumps(col_map, ensure_ascii=False), "success"))
    batch_id = cur2.lastrowid
    for c in cases:
        cur3 = conn.execute(
            "INSERT INTO case_item (case_batch_id,row_index,case_id,case_step,raw_row_json) VALUES (?,?,?,?,?)",
            (batch_id, c["row_index"], c["case_id"], c["case_step"], json.dumps(c["raw_row"], ensure_ascii=False)))
        c["db_id"] = cur3.lastrowid
    conn.commit(); conn.close()

    sid = f"case_{uuid.uuid4().hex}"
    set_session(sid, {"case_batch_id": batch_id, "uploaded_file_id": uf_id,
                      "cases": cases, "column_mapping": col_map,
                      "file_bytes": file_bytes, "filename": filename})

    # 写缓存文件
    cache_dir = os.path.join(UPLOAD_DIR, "cache")
    os.makedirs(cache_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(filename)[0]
    cache_path = os.path.join(cache_dir, f"cases_{base}_{ts}_{fhash[:8]}.json")
    try:
        with open(cache_path, "w", encoding="utf-8") as cf:
            json.dump({"meta": {"session_id": sid, "filename": filename, "file_hash": fhash,
                                "case_count": len(cases), "column_mapping": col_map,
                                "sheet_names": parsed.get("sheet_names", []),
                                "parsed_at": datetime.now().isoformat()},
                       "cases": [{"row_index": c["row_index"], "case_id": c["case_id"],
                                  "case_step": c["case_step"], "raw_row": c["raw_row"]}
                                 for c in cases]}, cf, ensure_ascii=False, indent=2)
        logger.info("Case cache written: %s", cache_path)
    except Exception as e:
        logger.warning("Failed to write case cache: %s", e)

    return jsonify({
        "case_session_id": sid,
        "case_count": len(cases),
        "sheet_names": parsed.get("sheet_names",[]),
        "column_mapping": col_map,
        "cases_preview": [{"row_index": c["row_index"], "case_id": c["case_id"], "case_step": c["case_step"]}
                           for c in cases[:20]],
    })

# ── Match ──────────────────────────────────────────────────────────────────────
@app.route("/api/match/run", methods=["POST"])
def api_run_match():
    body = request.get_json(force=True) or {}
    sig_sid = body.get("signal_session_id")
    case_sid = body.get("case_session_id")
    if not sig_sid or not case_sid:
        return err("缺少 signal_session_id 或 case_session_id")
    sig_sess = get_session(sig_sid)
    if not sig_sess: return err("信号会话不存在，请重新上传信号文件", 404)
    case_sess = get_session(case_sid)
    if not case_sess: return err("用例会话不存在，请重新上传测试用例文件", 404)

    mc = body.get("model_config", {})
    task_code = f"task_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO match_task (task_code,status,signal_file_id,case_file_id,model_name,model_base_url,temperature)
           VALUES (?,?,?,?,?,?,?)""",
        (task_code, "pending", sig_sess.get("uploaded_file_id"), case_sess.get("uploaded_file_id"),
         mc.get("model", DEEPSEEK_MODEL), mc.get("base_url", DEEPSEEK_BASE_URL), mc.get("temperature", 0)))
    task_id = cur.lastrowid
    conn.commit(); conn.close()

    logger.info(f"▶ Starting match task {task_code} | cases={len(case_sess['cases'])} signals={len(sig_sess['flat_signals'])}")
    results = run_match(task_id, sig_sess["flat_signals"], case_sess["cases"], mc, task_code)

    matched = sum(1 for r in results if r.get("matched"))
    return jsonify({
        "task_id": task_code,
        "status": "success",
        "total": len(results),
        "matched_count": matched,
        "unmatched_count": len(results) - matched,
        "results": results,
    })

# ── Export ─────────────────────────────────────────────────────────────────────
@app.route("/api/export/fill", methods=["POST"])
def api_export():
    body = request.get_json(force=True) or {}
    task_code = body.get("task_id")
    case_sid = body.get("case_session_id")
    if not task_code or not case_sid:
        return err("需要提供 task_id 和 case_session_id")
    conn = get_db()
    row = conn.execute("SELECT id FROM match_task WHERE task_code=?", (task_code,)).fetchone()
    if not row: return err(f"任务不存在: {task_code}", 404)
    task_id = row["id"]
    results_rows = conn.execute("SELECT result_json FROM case_match_result WHERE task_id=?", (task_id,)).fetchall()
    conn.close()
    results = []
    for r in results_rows:
        try: results.append(json.loads(r["result_json"]))
        except: pass
    if not results: return err("该任务暂无匹配结果", 404)

    case_sess = get_session(case_sid)
    if not case_sess: return err("用例会话不存在，请重新上传", 404)
    col_map = case_sess.get("column_mapping", {})
    case_id_col = col_map.get("case_id", "用例编号")
    try:
        data = export_excel_bytes(case_sess["file_bytes"], case_sess["filename"], results, case_id_col)
    except Exception as e:
        return err(f"导出失败: {e}", 500)
    return Response(data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{task_code}_filled.xlsx"'})

# ── Prompt Preview ─────────────────────────────────────────────────────────────
@app.route("/api/prompts/preview", methods=["POST"])
def api_prompt_preview():
    body = request.get_json(force=True) or {}
    case_step = body.get("case_step","")
    sig_sid = body.get("signal_session_id","")
    sig_sess = get_session(sig_sid)
    if not sig_sess: return err("信号会话不存在", 404)
    sem = normalize_case(case_step)
    cands = recall_candidates(case_step, sig_sess["flat_signals"], sem)
    prompts = build_prompts("preview", case_step, sem, cands)
    return jsonify({"system_prompt": prompts["system_prompt"], "user_prompt": prompts["user_prompt"],
                    "candidate_count": len(cands)})

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 8000))
    logger.info(f"🚀 Starting server on http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
